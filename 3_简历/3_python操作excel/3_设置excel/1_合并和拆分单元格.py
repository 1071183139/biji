import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors

wb = openpyxl.Workbook()

# Font 字体
# ws = wb.active
# ws.title = 'Font'
#
# # default 11pt, Calibri
# italic24Font = Font(size=24, italic=True)
# ws['B3'].font = italic24Font
# ws['B3'] = '24pt Italic'
#
# boldRedFont = Font(name='Times New Roman', bold=True, color=colors.RED)
# ws['A1'].font = boldRedFont
# ws['A1'] = 'Bold Red Times New Roman'
#
# # Formulas 公式
# ws = wb.create_sheet('Formula')
# ws['A1'] = 200
# ws['A2'] = 300
# ws['A3'] = '=SUM(A1:A2)'
#
# # Setting row height and column width 设置宽和高
# ws = wb.create_sheet('dimensions')
# ws['A1'] = 'Tall row'
# ws.row_dimensions[1].height = 70
# ws['B2'] = 'Wide column'
# ws.column_dimensions['B'].width = 20

# Merging cells 合并单元格
ws = wb.create_sheet('merged')
ws.merge_cells('A1:D3')
ws['A1'] = 'Twelve cells merged together'
ws.merge_cells('C5:D5')
ws['C5'] = 'Two merged cells'
#
# # Unmerging cells 拆分单元格
ws = wb.copy_worksheet(wb.get_sheet_by_name('merged'))
ws.title = 'unmerged'
ws.unmerge_cells('A1:D3')
ws.unmerge_cells('C5:D5')

wb.save('style.xlsx')