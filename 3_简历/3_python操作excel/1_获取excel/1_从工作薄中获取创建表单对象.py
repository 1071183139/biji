import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('example.xlsx')

# 从工作薄中获取一个表单(sheet)对象
sheets = wb.sheetnames
print(sheets, type(sheets))

# 创建一个表单
mySheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

# 获取指定的表单
sheet3 = wb.get_sheet_by_name('Sheet3')
sheet4 = wb['mySheet']
for sheet in wb:
    print(sheet.title)
