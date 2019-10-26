import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('example.xlsx')
ws = wb.active  # 当前活跃的表单

# 把数字转换成字母
print(get_column_letter(3), get_column_letter(47), get_column_letter(900))

# 把字母转换成数字
print(column_index_from_string('AAH'))