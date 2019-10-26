import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('example.xlsx')
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单
print(ws)
print(ws['A1'])  # 获取A列的第一个对象
print(ws['A1'].value)

c = ws['B1']
print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))  # 打印这个单元格对象所在的行列的数值和内容
print('Cell {} is {}\n'.format(c.coordinate, c.value))  # 获取单元格对象的所在列的行数和值