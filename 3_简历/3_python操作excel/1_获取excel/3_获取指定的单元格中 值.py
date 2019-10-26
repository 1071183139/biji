import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('example.xlsx')
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单

print(ws.cell(row=1, column=2))  # 获取第一行第二列的单元格
print(ws.cell(row=1, column=2).value)
for i in range(1, 8, 2):  # 获取1,3,5,7 行第二列的值
    print(i, ws.cell(row=i, column=2).value)