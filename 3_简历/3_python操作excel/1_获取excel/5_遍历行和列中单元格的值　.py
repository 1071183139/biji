import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('example.xlsx')
ws = wb.active  # 当前活跃的表单

col_range = ws['B:C']
row_range = ws[2:6]

# for col in col_range:  # 打印BC两列单元格中的值内容
#     for cell in col:
#         print(cell.value)

# for row in row_range:  # 打印 2-5行中所有单元格中的值
#     for cell in row:
#         print(cell.value)
#
for row in ws.iter_rows(min_row=1, max_row=2, max_col=3):  # 打印1-2行，1-3列中的内容
    for cell in row:
        print(cell.value)