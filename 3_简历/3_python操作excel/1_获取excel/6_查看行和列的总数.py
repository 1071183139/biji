import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('example.xlsx')
ws = wb.active  # 当前活跃的表单

print('{}行 {}列'.format(ws.max_row, ws.max_column))