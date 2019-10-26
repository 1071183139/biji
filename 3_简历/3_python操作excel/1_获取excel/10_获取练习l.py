import openpyxl
# # 获取工作簿对象
wb=openpyxl.load_workbook('example.xlsx')
#
# #创建表单
# mySheet = wb.create_sheet('mySheet')
#
# #获取所有娿表单
# print(wb.sheetnames)
#
# #获取指定的表单
# sheet3 = wb.get_sheet_by_name('mySheet')
# print(sheet3)
#
# sheet4=wb['mySheet']
# print(sheet4)


# 获取当前活跃的表单对象
ws=wb.active
# print(ws)

# 获取指定单元格信息
# print(ws['A1'].value)

# 获取第一行第二列的单元格
# print(ws.cell(row=1,column=2).value)

# 获取1,3,5,7 行第二列的值
# for i in range(1,8,2):
#     print(i,ws.cell(row=i,column=2).value)

# 获取整个C列
# print(ws['C'])
# c=ws['C']
# for i in c:
#     print(i.value)

# 获取第一行
# print(ws[1])


# 打印BC两列单元格中的值内容
# col_rang=ws['B:C']
# for col in col_rang:
#     for item in col:
#         print(item.value)

# 打印 2-5行中所有单元格中的值
# row_ran=ws[2:5]
# for row in row_ran:
#     for item in row:
#         print


# 查看行和列的总数
# ws1=wb['Sheet2']
# print(ws1.max_row,ws1.max_column)

