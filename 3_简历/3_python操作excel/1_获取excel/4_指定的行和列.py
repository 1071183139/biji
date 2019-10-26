import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('example.xlsx')
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单

# 从表单中获取行和列

colC = ws['C']  # 获取整个C列
print(colC)
row6 = ws[6]  # 获取第6行
print(row6, type(row6))