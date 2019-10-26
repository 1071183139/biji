import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active

wb.create_sheet(index=0, title='First Sheet')
wb.create_sheet(index=1, title='Middle Sheet')
print(wb.get_sheet_names())  # 获取当前工作薄的名字

wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
print(wb.get_sheet_names())

# wb.save('temp1.xlsx')